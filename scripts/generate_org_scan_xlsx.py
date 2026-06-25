"""Generate GitHub org scan XLSX from scan output data."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    import sys

    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

ROWS = """metR-Frontend|dev|YES|NO
metR-Frontend|devops-testing|YES|NO
metR-Frontend|main|NO|NO
api-metapercept|dev|N/A|BRANCH NOT FOUND
api-metapercept|devops-testing|N/A|BRANCH NOT FOUND
api-metapercept|main|NO|NO
metapercept-main|dev|N/A|BRANCH NOT FOUND
metapercept-main|devops-testing|N/A|BRANCH NOT FOUND
metapercept-main|main|NO|NO
metR-API-Gateway|dev|YES|NO
metR-API-Gateway|devops-testing|YES|NO
metR-API-Gateway|main|NO|NO
metR-Converter-PDF|dev|N/A|BRANCH NOT FOUND
metR-Converter-PDF|devops-testing|N/A|BRANCH NOT FOUND
metR-Converter-PDF|main|NO|NO
blog-metapercept|dev|N/A|BRANCH NOT FOUND
blog-metapercept|devops-testing|N/A|BRANCH NOT FOUND
blog-metapercept|main|NO|NO
Asish_Test_Repo|dev|N/A|BRANCH NOT FOUND
Asish_Test_Repo|devops-testing|N/A|BRANCH NOT FOUND
Asish_Test_Repo|main|NO|NO
metR-Converter-HTML|dev|YES|NO
metR-Converter-HTML|devops-testing|YES|NO
metR-Converter-HTML|main|NO|NO
metR-Converter-Markdown|dev|YES|NO
metR-Converter-Markdown|devops-testing|YES|NO
metR-Converter-Markdown|main|NO|NO
MetR-Demo|dev|N/A|BRANCH NOT FOUND
MetR-Demo|devops-testing|N/A|BRANCH NOT FOUND
MetR-Demo|main|NO|NO
metR-Client-Admin|dev|N/A|BRANCH NOT FOUND
metR-Client-Admin|devops-testing|N/A|BRANCH NOT FOUND
metR-Client-Admin|main|N/A|BRANCH NOT FOUND
metR-Converter-DOCX|dev|YES|NO
metR-Converter-DOCX|devops-testing|YES|NO
metR-Converter-DOCX|main|NO|NO
metR-Server-Admin|dev|NO|NO
metR-Server-Admin|devops-testing|NO|NO
metR-Server-Admin|main|N/A|BRANCH NOT FOUND
metR-Converter-FrameMaker|dev|YES|NO
metR-Converter-FrameMaker|devops-testing|YES|NO
metR-Converter-FrameMaker|main|NO|NO
metR_Determinacy_Engine|dev|N/A|BRANCH NOT FOUND
metR_Determinacy_Engine|devops-testing|N/A|BRANCH NOT FOUND
metR_Determinacy_Engine|main|NO|NO
metR-Phase2-shared-pipelines|dev|N/A|BRANCH NOT FOUND
metR-Phase2-shared-pipelines|devops-testing|N/A|BRANCH NOT FOUND
metR-Phase2-shared-pipelines|main|YES|NO
admin-metapercept|dev|N/A|BRANCH NOT FOUND
admin-metapercept|devops-testing|N/A|BRANCH NOT FOUND
admin-metapercept|main|NO|NO
portfolio-metapercept|dev|N/A|BRANCH NOT FOUND
portfolio-metapercept|devops-testing|N/A|BRANCH NOT FOUND
portfolio-metapercept|main|NO|NO
careers-metapercept|dev|N/A|BRANCH NOT FOUND
careers-metapercept|devops-testing|N/A|BRANCH NOT FOUND
careers-metapercept|main|NO|NO
training-metapercept|dev|N/A|BRANCH NOT FOUND
training-metapercept|devops-testing|N/A|BRANCH NOT FOUND
training-metapercept|main|N/A|BRANCH NOT FOUND
training-appadmin-metapercept|dev|N/A|BRANCH NOT FOUND
training-appadmin-metapercept|devops-testing|N/A|BRANCH NOT FOUND
training-appadmin-metapercept|main|N/A|BRANCH NOT FOUND
portfolio-adminpanel-metapercept|dev|N/A|BRANCH NOT FOUND
portfolio-adminpanel-metapercept|devops-testing|N/A|BRANCH NOT FOUND
portfolio-adminpanel-metapercept|main|N/A|BRANCH NOT FOUND
careers-adminpanel-metapercept|dev|N/A|BRANCH NOT FOUND
careers-adminpanel-metapercept|devops-testing|N/A|BRANCH NOT FOUND
careers-adminpanel-metapercept|main|N/A|BRANCH NOT FOUND
blog-adminpanel-metapercept|dev|N/A|BRANCH NOT FOUND
blog-adminpanel-metapercept|devops-testing|N/A|BRANCH NOT FOUND
blog-adminpanel-metapercept|main|N/A|BRANCH NOT FOUND
metR-DocManager|dev|N/A|BRANCH NOT FOUND
metR-DocManager|devops-testing|N/A|BRANCH NOT FOUND
metR-DocManager|main|NO|NO
ai-devops-teammate|dev|N/A|BRANCH NOT FOUND
ai-devops-teammate|devops-testing|N/A|BRANCH NOT FOUND
ai-devops-teammate|main|NO|NO
Kajal_Test_Repo|dev|N/A|BRANCH NOT FOUND
Kajal_Test_Repo|devops-testing|N/A|BRANCH NOT FOUND
Kajal_Test_Repo|main|NO|NO
KUBERNETES|dev|N/A|BRANCH NOT FOUND
KUBERNETES|devops-testing|N/A|BRANCH NOT FOUND
KUBERNETES|main|NO|NO
github-org-mcp-server|dev|N/A|BRANCH NOT FOUND
github-org-mcp-server|devops-testing|N/A|BRANCH NOT FOUND
github-org-mcp-server|main|NO|NO
WORKDAY-SCHEDULER-SETUP|dev|N/A|BRANCH NOT FOUND
WORKDAY-SCHEDULER-SETUP|devops-testing|N/A|BRANCH NOT FOUND
WORKDAY-SCHEDULER-SETUP|main|NO|NO
oxygen-web-author|dev|N/A|BRANCH NOT FOUND
oxygen-web-author|devops-testing|N/A|BRANCH NOT FOUND
oxygen-web-author|main|NO|NO
AWS-BEDROCK-SAM|dev|N/A|BRANCH NOT FOUND
AWS-BEDROCK-SAM|devops-testing|N/A|BRANCH NOT FOUND
AWS-BEDROCK-SAM|main|NO|NO
metR-Phase2|dev|N/A|BRANCH NOT FOUND
metR-Phase2|devops-testing|N/A|BRANCH NOT FOUND
metR-Phase2|main|NO|NO
RuleSet-Branches-Secure|dev|N/A|BRANCH NOT FOUND
RuleSet-Branches-Secure|devops-testing|N/A|BRANCH NOT FOUND
RuleSet-Branches-Secure|main|NO|NO
MetR-Master-AI-Agent|dev|N/A|BRANCH NOT FOUND
MetR-Master-AI-Agent|devops-testing|N/A|BRANCH NOT FOUND
MetR-Master-AI-Agent|main|N/A|BRANCH NOT FOUND
dita2-validation-transformation|dev|N/A|BRANCH NOT FOUND
dita2-validation-transformation|devops-testing|N/A|BRANCH NOT FOUND
dita2-validation-transformation|main|NO|NO
EnableX-Sourcing-files|dev|N/A|BRANCH NOT FOUND
EnableX-Sourcing-files|devops-testing|N/A|BRANCH NOT FOUND
EnableX-Sourcing-files|main|NO|NO
Nirmata-Sourcing-files|dev|N/A|BRANCH NOT FOUND
Nirmata-Sourcing-files|devops-testing|N/A|BRANCH NOT FOUND
Nirmata-Sourcing-files|main|NO|NO
demo-repository|dev|N/A|BRANCH NOT FOUND
demo-repository|devops-testing|N/A|BRANCH NOT FOUND
demo-repository|main|NO|NO
metR-DedUp|dev|N/A|BRANCH NOT FOUND
metR-DedUp|devops-testing|N/A|BRANCH NOT FOUND
metR-DedUp|main|NO|NO
metR-Converter-SitecoreXML|dev|N/A|BRANCH NOT FOUND
metR-Converter-SitecoreXML|devops-testing|N/A|BRANCH NOT FOUND
metR-Converter-SitecoreXML|main|NO|NO
metR-Converter-PropDITA|dev|N/A|BRANCH NOT FOUND
metR-Converter-PropDITA|devops-testing|N/A|BRANCH NOT FOUND
metR-Converter-PropDITA|main|NO|NO
metR-Converter-DocBook|dev|N/A|BRANCH NOT FOUND
metR-Converter-DocBook|devops-testing|N/A|BRANCH NOT FOUND
metR-Converter-DocBook|main|NO|NO
metR-Core-Infrastructure|dev|N/A|BRANCH NOT FOUND
metR-Core-Infrastructure|devops-testing|N/A|BRANCH NOT FOUND
metR-Core-Infrastructure|main|NO|NO
metR-InfinityDocSite|dev|N/A|BRANCH NOT FOUND
metR-InfinityDocSite|devops-testing|N/A|BRANCH NOT FOUND
metR-InfinityDocSite|main|NO|NO
metR-Output-XML-JSON-AI|dev|N/A|BRANCH NOT FOUND
metR-Output-XML-JSON-AI|devops-testing|N/A|BRANCH NOT FOUND
metR-Output-XML-JSON-AI|main|NO|NO
metR-Output-HTML5-CMS|dev|N/A|BRANCH NOT FOUND
metR-Output-HTML5-CMS|devops-testing|N/A|BRANCH NOT FOUND
metR-Output-HTML5-CMS|main|NO|NO
metR-Output-PDF|dev|N/A|BRANCH NOT FOUND
metR-Output-PDF|devops-testing|N/A|BRANCH NOT FOUND
metR-Output-PDF|main|NO|NO
metR-DocPublisher|dev|N/A|BRANCH NOT FOUND
metR-DocPublisher|devops-testing|N/A|BRANCH NOT FOUND
metR-DocPublisher|main|NO|NO
metR-DocEditor|dev|N/A|BRANCH NOT FOUND
metR-DocEditor|devops-testing|N/A|BRANCH NOT FOUND
metR-DocEditor|main|NO|NO
metR-DocMigrate|dev|N/A|BRANCH NOT FOUND
metR-DocMigrate|devops-testing|N/A|BRANCH NOT FOUND
metR-DocMigrate|main|NO|NO
metR-XSLT-Library|dev|N/A|BRANCH NOT FOUND
metR-XSLT-Library|devops-testing|N/A|BRANCH NOT FOUND
metR-XSLT-Library|main|NO|NO"""


def parse_rows():
    records = []
    for line in ROWS.strip().splitlines():
        repo, branch, codeowners, protection = line.split("|")
        records.append(
            {
                "Repository": repo,
                "Branch": branch,
                "CODEOWNERS": codeowners,
                "BranchProtection": protection,
            }
        )
    return records


def risk_level(row):
    if row["BranchProtection"] != "NO":
        return "N/A"
    if row["CODEOWNERS"] == "YES":
        return "Medium — CODEOWNERS only"
    if row["CODEOWNERS"] == "NO":
        return "High — no CODEOWNERS"
    return "Low — branch missing"


def style_header(ws, row=1):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[letter].width = min(width, 45)


def build_summary(records):
    repos = sorted({r["Repository"] for r in records})
    summary = []
    for repo in repos:
        repo_rows = [r for r in records if r["Repository"] == repo]
        branches_found = [r["Branch"] for r in repo_rows if r["BranchProtection"] != "BRANCH NOT FOUND"]
        codeowners_yes = sum(1 for r in repo_rows if r["CODEOWNERS"] == "YES")
        protection_yes = sum(1 for r in repo_rows if r["BranchProtection"] == "YES")
        summary.append(
            {
                "Repository": repo,
                "Branches Scanned": len(repo_rows),
                "Branches Found": len(branches_found),
                "CODEOWNERS (YES count)": codeowners_yes,
                "Branch Protection (YES count)": protection_yes,
                "Gap": "Needs branch protection" if protection_yes == 0 and branches_found else "",
            }
        )
    return summary


def main():
    records = parse_rows()
    out = Path(__file__).resolve().parents[1] / "reports" / "github-org-scan.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: Full scan
    ws = wb.active
    ws.title = "Org Scan"
    headers = ["Repository", "Branch", "CODEOWNERS", "BranchProtection", "Risk Notes"]
    ws.append(headers)
    for row in records:
        ws.append(
            [
                row["Repository"],
                row["Branch"],
                row["CODEOWNERS"],
                row["BranchProtection"],
                risk_level(row),
            ]
        )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    style_header(ws)
    autosize_columns(ws)

    # Sheet 2: Summary by repository
    ws2 = wb.create_sheet("Summary")
    summary_headers = [
        "Repository",
        "Branches Scanned",
        "Branches Found",
        "CODEOWNERS (YES count)",
        "Branch Protection (YES count)",
        "Gap",
    ]
    ws2.append(summary_headers)
    for item in build_summary(records):
        ws2.append([item[h] for h in summary_headers])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    style_header(ws2)
    autosize_columns(ws2)

    # Sheet 3: Gaps — main branch without protection
    ws3 = wb.create_sheet("Gaps - Main Branch")
    gap_headers = ["Repository", "Branch", "CODEOWNERS", "BranchProtection", "Action Required"]
    ws3.append(gap_headers)
    for row in records:
        if row["Branch"] == "main" and row["BranchProtection"] == "NO":
            action = (
                "Add CODEOWNERS + enable branch protection"
                if row["CODEOWNERS"] == "NO"
                else "Enable branch protection rules"
            )
            ws3.append([row["Repository"], row["Branch"], row["CODEOWNERS"], row["BranchProtection"], action])
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    style_header(ws3)
    autosize_columns(ws3)

    wb.save(out)
    print(f"Created: {out}")
    print(f"Rows: {len(records)} | Repositories: {len({r['Repository'] for r in records})}")


if __name__ == "__main__":
    main()
